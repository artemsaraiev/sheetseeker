# Algo description is in the tutorial/ folder


from flask import Flask, render_template, request, redirect, url_for, flash, session
import threading
import os
import io
import csv
from dotenv import load_dotenv
import openpyxl
from openai import OpenAI
from pinecone import Pinecone
# from create_responses import send_message_to_llm
import re
import concurrent.futures
from ratelimit import limits, sleep_and_retry
from flask_socketio import SocketIO, emit

from utils.data_formatting import json_to_string, create_filtered_meta_excel
from utils.gpt_call import send_gpt_request
from utils.embeds import generate_embeddings, upsert_chunks_query
from utils.agents import FormatQuestionCall, AnswerCall

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads/'
load_dotenv()
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY')
app.config['EMBED_API_KEY'] = os.environ.get('EMBED_API_KEY')
app.config['PINECONE_API_KEY'] = os.environ.get('PINECONE_API_KEY')
app.config['OPENROUTER_API_KEY'] = os.environ.get('OPENROUTER_API_KEY')
app.config['EMBED_MODEL'] = os.environ.get('EMBED_MODEL')
app.config['BASE_MODEL'] = os.environ.get('BASE_MODEL')
app.config['INDEX_NAME'] = os.environ.get('INDEX_NAME')
NAMESPACE = ""

pc = Pinecone(api_key=app.config['PINECONE_API_KEY'])
INDEX = pc.Index(app.config['INDEX_NAME'])
INDEX_DIMENSION = 1536


def clear_namespace(namespace, index = INDEX):
    # just add a single vector to the namespace, because 
    # deleting non-existing namespace will raise an error
    index.upsert(
    vectors=[
        {"id": "A", "values": [0.1]*INDEX_DIMENSION},
    ],
    namespace=namespace,
    )
    index.delete(namespace = namespace, delete_all = True)

def find_column_name(sheet, column_index):
    # Go upward in the column until a non-numeric cell (a header) is found
    current_row = 0
    while current_row < sheet.max_row and isinstance(sheet.cell(row=current_row+1, column=column_index+1).value, (int, float)):
        current_row += 1
    return sheet.cell(row=current_row+1, column=column_index+1).value


def get_formatted_quiries(user_query):
    filtered_query = FormatQuestionCall(user_query)
    filtered_query.respond()
    queries_formatted =  ({"text_to_embed": query, "id": str(ix+1)} for ix, query in enumerate(filtered_query.queries))
    print(f'{filtered_query.queries = }')
    return queries_formatted, filtered_query

def embed_queries(queries_formatted):
    new_namespace = "QUERIES"
    clear_namespace(new_namespace, INDEX)
    upsert_chunks_query(queries_formatted, "QUERIES")

def get_similarities(query, namespace, index = INDEX, top_k=500, filter = None):
    temp_client = OpenAI(
    api_key=os.environ.get('EMBED_API_KEY'),
    )

    description_response = temp_client.embeddings.create(
            input=query,
            model=app.config['EMBED_MODEL'],
        )
    query_embed_vec = description_response.data[0].embedding
    if filter is None:
        query = index.query(
            namespace=namespace,
            vector=query_embed_vec,
            top_k=top_k,
            include_metadata=True
        )
    else:
        query = index.query(
            namespace=namespace,
            vector=query_embed_vec,
            filter = filter,
            top_k=top_k,
            include_metadata=True
        )
    return query

def answer_single_query(query, table_path):
        similarities = get_similarities(query, table_path, INDEX)

        # create filtered table
        # cell_set = create_similar_cell_set(similarities, threshold=0.2)
        filtered_path = f"filtered_table.xlsx"
        print(f'filtered excel for {table_path} is being created at {filtered_path}')

        create_filtered_meta_excel(app.config['UPLOAD_FOLDER'] + table_path, filtered_path, similarities)
        # create_filtered_excel(table_path, filtered_path, cell_set)

        answer_call = AnswerCall(query)
        answer_call.load_table(filtered_path)
        answer_call.respond()

        query_result = answer_call.highlighted_json
        return query_result

def process_user_query(user_query, table_path):
    queries_formatted, filtered_query = get_formatted_quiries(user_query)
    embed_queries(queries_formatted)

    query_answers_list = []

    for query in filtered_query.queries:
        query_result = answer_single_query(query, table_path)
        query_answers_list.append(json_to_string(query_result))
       
    return query_answers_list

def send_message_to_llm(query, namespace):
    # getting relevant embeddings first
    query_answer_list = process_user_query(query, namespace)
    respond_massage = '\n'.join(json_to_string(string) for string in query_answer_list)

    return respond_massage

@app.route('/clear_history', methods=['POST'])
def clear_history():
    if NAMESPACE in session['chat_history']:
        session['chat_history'][NAMESPACE] = []
    session.modified = True
    return redirect(url_for('chat'))


@app.route('/', methods=['GET', 'POST'])
def home():
    session['chat_history'] = {}

    if request.method == 'POST':
        selected_name = request.form.get('names')
        if selected_name:
            global NAMESPACE
            NAMESPACE = selected_name
            session['chat_history'][NAMESPACE] = []
        return redirect(url_for('chat'))

    return render_template('index.html')


@app.route('/send_message', methods=['POST'])
def send_message():
    text = request.form.get('message')
    if text:
        global NAMESPACE
        print("send mes", NAMESPACE)
        session['chat_history'][NAMESPACE].append(
            {"role": "user", "content": text})
        answer = send_message_to_llm(text, NAMESPACE)
        print("answer", answer)
        print(session['chat_history'])
        session['chat_history'][NAMESPACE].append(
            {"role": "assistant", "content": answer})
        session.modified = True
    return redirect(url_for('chat'))


@app.route('/chat')
def chat():
    # Handle chat history display based on namespace
    global NAMESPACE
    chat_history = session['chat_history'][NAMESPACE]
    return render_template('chat.html', chat_history=chat_history, namespace=NAMESPACE)


@app.route('/upload', methods=['POST'])
def upload_file():
    file = request.files.get('file')
    if file and file.filename:
        filename = file.filename
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        session['uploaded_file'] = filepath
        global NAMESPACE
        NAMESPACE = filename
        generate_embeddings(filepath, filename)
        chat_history = session['chat_history'][NAMESPACE] = []
        print("in upload", chat_history)
        return redirect(url_for('chat'))
    else:
        flash('No file selected.', 'error')
        return redirect(url_for('home'))


if __name__ == '__main__':
    app.run(debug=True)
