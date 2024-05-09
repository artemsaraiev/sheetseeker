from flask import Flask, render_template, request, redirect, url_for, flash, session
import threading
import os
import io
import csv
from dotenv import load_dotenv
import openpyxl
from openai import OpenAI
from pinecone import Pinecone
import re
import concurrent.futures
from ratelimit import limits, sleep_and_retry
from flask_socketio import SocketIO, emit

from .gpt_call import send_gpt_request
from .data_formatting import json_to_string, create_filtered_excel, create_similar_cell_set
from .agents import FormatQuestionCall, AnswerCall

pc = Pinecone(api_key=os.environ.get('PINECONE_API_KEY'))
INDEX = pc.Index(os.environ.get('INDEX_NAME'))


def generate_fin_term_embeds(file_path):
    def get_fin_terms_cell():
        # Open the workbook with data_only=True to get the values instead of formulas
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        terms = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]

            # Iterate over all cells in the current sheet
            for row in sheet.iter_rows():
                for term in row:
                    common_data = {
                        "sheet": sheet_name,
                        "row": term.row,
                        "col": term.column,
                        "window_width": 0,
                        "window_height": 0,
                    }
                    if term.value is None or isinstance(term.value, str) and not term.value.strip():
                        continue
                    text_to_embed = str(term.value)
                    # print(f'{text_to_embed = }')
                    if not text_to_embed.isnumeric():
                        common_data.update({"text_to_embed": text_to_embed})
                        common_data.update({"type": "term"})
                    elif text_to_embed.isnumeric():
                        # Find the first text cell to the left
                        left_cell = term
                        left_text = ""
                        while left_cell.column > 1:
                            left_cell = sheet.cell(
                                row=left_cell.row, column=left_cell.column - 1)
                            if isinstance(left_cell.value, str) and not left_cell.value.isnumeric():
                                left_text = left_cell.value
                                break

                        up_cell = term
                        up_text = ""
                        while up_cell.row > 1:
                            up_cell = sheet.cell(
                                row=up_cell.row - 1, column=up_cell.column)
                            if isinstance(up_cell.value, str):
                                up_text = up_cell.value
                                break
                        text_to_embed = left_text + " " + up_text
                        # print(f'Combined text to embed: {text_to_embed}')
                        common_data.update({"text_to_embed": text_to_embed})
                        common_data.update({"type": "value"})
                    terms.append(common_data)
        return terms
    return get_fin_terms_cell()


def generate_embeddings(file_path, filename):
    fin_term_embeds = generate_fin_term_embeds(file_path)
    # sliding_window_embeds = generate_sliding_window_emeds(file_path)
    combined_terms = fin_term_embeds #+ sliding_window_embeds
    chunk_id = 1
    for term in combined_terms:
        term['id'] = str(chunk_id)
        chunk_id += 1

    # let file_path be namespace name
    upsert_chunks(combined_terms, str(filename))


def upsert_chunks_query(terms, namespace, index=INDEX):
    embeds = create_vectors(terms)
    new_embeds = []
    for emb in embeds:
        new_embeds.append({"id": emb["id"],
                           "values": emb["values"],
                           })

    chunks = [new_embeds[i:i + 100] for i in range(0, len(new_embeds), 100)]
    for chunk in chunks:
        index.upsert(vectors=chunk, namespace=namespace)


def create_vectors(terms):
    temp_client = OpenAI(api_key=os.environ.get('EMBED_API_KEY'))

    @sleep_and_retry
    @limits(calls=40, period=1)
    def get_embedding(obj):
        response = temp_client.embeddings.create(
            input=obj['text_to_embed'],
            model="text-embedding-3-small",
        )
        embedding_vector = response.data[0].embedding
        obj['values'] = embedding_vector
        return obj

    with concurrent.futures.ThreadPoolExecutor() as executor:
        future_embeddings = [executor.submit(
            get_embedding, term) for term in terms]
        results = [future.result()
                   for future in concurrent.futures.as_completed(future_embeddings)]
    return results


def upsert_chunks(terms, namespace, index=INDEX):
    print("num of embedding objects", len(terms))
    print("creating embeddings")
    embeds = create_vectors(terms)
    print("done with embeddings")

    new_embeds = []
    # forming chunks to have format upsertable to pinecone
    for emb in embeds:
        new_embeds.append({"id": emb["id"], "values": emb["values"], "metadata": {
            "sheet": emb["sheet"], "row": emb["row"],
            "col": emb["col"], "window_width": emb["window_width"],
            "window_height": emb["window_height"], "type": emb["type"]}})

    chunks = [new_embeds[i:i + 100] for i in range(0, len(new_embeds), 100)]

    for chunk in chunks:
        # print("upserting chunk")
        index.upsert(vectors=chunk, namespace=namespace)


def convert_to_csv_string(data):
    si = io.StringIO()
    cw = csv.writer(si)
    cw.writerows(data)
    return si.getvalue()


def generate_sliding_window_emeds(file_path):
    snapshots = get_xlsx_snapshots(file_path)

    def process_snapshot(snapshot):
        return send_gpt_request([{"role": "user", "content": snapshot['text']}], "openai/gpt-3.5-turbo", "generate_questions")

    with concurrent.futures.ThreadPoolExecutor() as executor:
        futures = [executor.submit(process_snapshot, snapshot)
                   for snapshot in snapshots]
        for snapshot, future in zip(snapshots, concurrent.futures.as_completed(futures)):
            snapshot['text_to_embed'] = future.result()

    return snapshots


def get_xlsx_snapshots(file_path):
    wb = openpyxl.load_workbook(file_path)
    snapshots = []

    for sheet in wb:
        rows = list(sheet.iter_rows(values_only=True))
        num_rows = len(rows)
        if num_rows == 0:
            continue

        # Determine the number of windows and step size
        num_windows = min(15, num_rows)
        window_height = max(1, num_rows // num_windows)
        step_size = max(1, (num_rows - window_height) //
                        (num_windows - 1)) if num_windows > 1 else 1

        # Iterate over the rows with a sliding window
        for start in range(1, num_rows - window_height + 1, step_size):
            end = start + window_height
            window_data = rows[start:end]
            snapshot = {
                'sheet': sheet.title,
                'row': start,
                'col': 1,
                'type': "window",
                'window_width': len(window_data[0]),
                'window_height': len(window_data),
                'text': ''
            }
            headers = [find_column_name(sheet, i + 1)
                       for i in range(len(window_data[0]))]
            csv_data = [headers]
            for row in window_data:
                csv_data.append(list(row))

            csv_string = convert_to_csv_string(csv_data)
            snapshot['text'] = csv_string
            snapshots.append(snapshot)

    return snapshots


def find_column_name(sheet, column_index):
    current_row = 1
    while current_row < sheet.max_row and isinstance(sheet.cell(row=current_row+1, column=column_index).value, (int, float)):
        current_row += 1
    return sheet.cell(row=current_row+1, column=column_index).value
