import pandas as pd
import openpyxl
import re


def get_csv_data(spreadsheet_file_path) -> str:
    """
    This function reads a spreadsheet file and returns the raw data in CSV format.
    """
    result = "CSV DATA:\n\n"

    xl = pd.ExcelFile(spreadsheet_file_path)
    for sheet_name in xl.sheet_names:
        df = xl.parse(sheet_name, header=None)
        df.index += 1       # this 1-indexes the rows, so that the LLM returns the correct cell

        # outputting CSV data: we found a LLM prompt that works with CSV data
        result += f"Sheet Name: {sheet_name}\n{df.to_csv()}\n"

    return result

def create_similar_cell_set(similarities, threshold = 0.3):
    cell_list = set()
    for match in similarities['matches']:
        if match['score'] > threshold:
            cell = (match['metadata']['row'], match['metadata']['col'], match['metadata']['sheet'])
            cell_list.add(cell)

    return cell_list


def create_filtered_meta_excel(input_path, output_path, similarities, threshold = 0.4, window_size = 1):
    #update similarities with terms and scores
    cell_list = set()
    # Load the input workbook
    input_workbook = openpyxl.load_workbook(input_path)

    # add left and up terms of values to the set of cells
    for sheet_name in input_workbook.sheetnames:
        input_sheet = input_workbook[sheet_name]
        for match in similarities:
            # print(f'{match =}')
            if match['score'] > threshold:
                row, col, sheet = (int(match['metadata']['row']), 
                                   int(match['metadata']['col']), 
                                   match['metadata']['sheet'])
                if sheet != sheet_name:
                    continue
                cell = (row, col, sheet)
                cell_list.add(cell)

                if match['metadata']['type'] == "value":
                    for c_row in range(row-1, 0, -1):
                        left_term = input_sheet.cell(row=c_row, column=col).value
                        if isinstance(left_term, str):
                            if left_term != "":
                                cell_list.add((c_row, col, sheet))
                            break
                    for c_col in range(col-1, 0, -1):
                        up_term = input_sheet.cell(row=row, column=c_col).value
                        if isinstance(up_term, str):
                            if up_term != "":
                                cell_list.add((row, c_col, sheet))
                            break   
    
    print(len(cell_list))
    all_highlighted_cells = set()
    set_of_cells = cell_list.copy()
    # Create a new workbook for the output
    output_workbook = openpyxl.Workbook()

    # Iterate over each sheet in the input workbook
    for sheet_name in input_workbook.sheetnames:
        input_sheet = input_workbook[sheet_name]
        output_sheet = output_workbook.create_sheet(sheet_name)

        # Get the maximum row and column in the input sheet
        max_row = input_sheet.max_row
        max_col = input_sheet.max_column

        # Iterate over each row and column in the input sheet
        for cell in set_of_cells:
            row, col, sheet = cell
            if sheet != sheet_name:
                continue

            for r_off in range(-window_size, window_size+1): 
                for c_off in range(-window_size, window_size+1):
                    row_around, col_around = row + r_off, col + c_off
                    row_around = int(min(max_row, max(1, row_around)))
                    col_around = int(min(max_col, max(col_around, 1)))
                    if (row_around, col_around, sheet) not in all_highlighted_cells:
                        all_highlighted_cells.add((row_around, col_around, sheet))
                    output_sheet.cell(row=row_around, column=col_around).value = input_sheet.cell(row=row_around, column=col_around).value
            
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                if (row, col, sheet_name) not in all_highlighted_cells:
                    output_sheet.cell(row=row, column=col).value = None

    # Remove the default "Sheet" sheet from the output workbook
    output_workbook.remove(output_workbook["Sheet"])

    # Save the output workbook
    output_workbook.save(output_path)    

def create_filtered_excel(input_path, output_path, set_of_cells, window_size = 2):
    all_highlighted_cells = set_of_cells.copy()

    # Load the input workbook
    input_workbook = openpyxl.load_workbook(input_path)

    # Create a new workbook for the output
    output_workbook = openpyxl.Workbook()

    # Iterate over each sheet in the input workbook
    for sheet_name in input_workbook.sheetnames:
        input_sheet = input_workbook[sheet_name]
        output_sheet = output_workbook.create_sheet(sheet_name)

        # Get the maximum row and column in the input sheet
        max_row = input_sheet.max_row
        max_col = input_sheet.max_column

        # Iterate over each row and column in the input sheet
        for cell in set_of_cells:
            row, col, sheet = cell
            if sheet != sheet_name:
                continue

            for r_off in range(-window_size, window_size+1): 
                for c_off in range(-window_size, window_size+1):
                    row_around, col_around = row + r_off, col + c_off
                    row_around = int(min(max_row, max(1, row_around)))
                    col_around = int(min(max_col, max(col_around, 1)))
                    if (row_around, col_around, sheet) not in all_highlighted_cells:
                        all_highlighted_cells.add((row_around, col_around, sheet))
                    output_sheet.cell(row=row_around, column=col_around).value = input_sheet.cell(row=row_around, column=col_around).value
            
        for row in range(1, max_row + 1):
            for col in range(1, max_col + 1):
                if (row, col, sheet_name) not in all_highlighted_cells:
                    output_sheet.cell(row=row, column=col).value = None

    # Remove the default "Sheet" sheet from the output workbook
    output_workbook.remove(output_workbook["Sheet"])

    # Save the output workbook
    output_workbook.save(output_path)

def json_to_string(string):
    # Replace escape characters with their corresponding values
    string = string.replace("\\n", "\n")
    string = string.replace("\\'", "'")
    return string