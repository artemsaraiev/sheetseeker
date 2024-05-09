import os
import requests
import json
import pandas as pd
import sys
import openpyxl
import random
import itertools
import pandas as pd
from pinecone import Pinecone
from openpyxl import Workbook, load_workbook
from openai import OpenAI
import concurrent.futures

# !pip3 install python-dotenv
from dotenv import load_dotenv


import os
import os
import pandas as pd
from pinecone import Pinecone
from openpyxl import Workbook, load_workbook
from openai import OpenAI
from utils.data_formatting import create_filtered_meta_excel
from utils.agents import *


from dotenv import load_dotenv
# Get the current working directory
current_dir = os.getcwd()
# Construct the path to the .env file relative to the current directory
env_path = os.path.join(current_dir, '..', '.env')
# Load environment variables from .env file
load_dotenv(env_path)

SECRET_KEY = os.environ.get('SECRET_KEY')
EMBED_API_KEY = os.environ.get('EMBED_API_KEY')
PINECONE_API_KEY = os.environ.get('PINECONE_API_KEY')
OPENROUTER_API_KEY = os.environ.get('OPENROUTER_API_KEY')
EMBED_MODEL = os.environ.get('EMBED_MODEL')
BASE_MODEL = os.environ.get('BASE_MODEL')
INDEX_NAME = os.environ.get('INDEX_NAME')

pc = Pinecone(api_key=PINECONE_API_KEY)
INDEX = pc.Index(INDEX_NAME)
INDEX_DIMENSION = 1536


def clear_namespace(namespace, index = INDEX):
    # just add a single vector to the namespace, because 
    # deleting non-existing namespace will raise an error
    index.upsert(
    vectors=[
        {"id": "A", "values": [0.1]*INDEX_DIMENSION},
    ],
    namespace=namespace,
    )
    index.delete(namespace = namespace, delete_all = True)

def create_embeddings(terms, embed_model=EMBED_MODEL):
    temp_client = OpenAI(api_key=EMBED_API_KEY)
    def get_embedding(obj):
        response = temp_client.embeddings.create(
            input=obj['text_to_embed'],
            model=embed_model,
        )
        embedding_vector = response.data[0].embedding
        obj['values'] = embedding_vector
        return obj

    with concurrent.futures.ThreadPoolExecutor() as executor:
        future_embeddings = [executor.submit(get_embedding, term) for term in terms]
        results = [future.result() for future in concurrent.futures.as_completed(future_embeddings)]
    return results

def upsert_chunks(terms, namespace, index=INDEX):
    embeds = create_embeddings(terms)
    new_embeds = []
    for emb in embeds:
        new_embeds.append({"id": emb["id"], 
                           "values": emb["values"], 
                           "metadata": {
                                "sheet": emb["sheet"],
                                "row": emb["row"],
                                "col": emb["col"],
                                "window_width": emb["window_width"],
                                "window_height": emb["window_height"],
                                "type": emb["type"]
                                }   
                        })

    chunks = [new_embeds[i:i + 100] for i in range(0, len(new_embeds), 100)]

    for chunk in chunks:
        index.upsert(vectors=chunk, namespace=namespace)

def upsert_chunks_query(queries, namespace, index=INDEX):
    embeds = create_embeddings(queries)
    new_embeds = []
    for emb in embeds:
        new_embeds.append({"id": emb["id"],
                           "values": emb["values"],   
                        })

    chunks = [new_embeds[i:i + 100] for i in range(0, len(new_embeds), 100)]

    for chunk in chunks:
        index.upsert(vectors=chunk, namespace=namespace)

def generate_fin_term_embeds(file_path):
    def get_fin_terms_cell():
        # Open the workbook with data_only=True to get the values instead of formulas
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        terms = []
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            
            # Iterate over all cells in the current sheet
            for row in sheet.iter_rows():
                for term in row:
                    common_data = {
                        "sheet": sheet_name,
                        "row": term.row,
                        "col": term.column,
                        "window_width": 0,
                        "window_height": 0,
                    }
                    if term.value is None or isinstance(term.value, str) and not term.value.strip():
                        continue
                    text_to_embed = str(term.value)
                    # print(f'{text_to_embed = }')
                    if not text_to_embed.isnumeric():
                        common_data.update({"text_to_embed": text_to_embed})
                        common_data.update({"type": "term"})
                    elif text_to_embed.isnumeric():
                        # Find the first text cell to the left
                        left_cell = term
                        left_text = ""
                        while left_cell.column > 1:
                            left_cell = sheet.cell(row=left_cell.row, column=left_cell.column - 1)
                            if isinstance(left_cell.value, str) and not left_cell.value.isnumeric():
                                left_text = left_cell.value
                                break

                        up_cell = term
                        up_text = ""
                        while up_cell.row > 1:
                            up_cell = sheet.cell(row=up_cell.row - 1, column=up_cell.column)
                            if isinstance(up_cell.value, str):
                                up_text = up_cell.value
                                break
                        text_to_embed = left_text + " " + up_text
                        # print(f'Combined text to embed: {text_to_embed}')
                        common_data.update({"text_to_embed": text_to_embed})
                        common_data.update({"type": "value"})
                    terms.append(common_data)
        return terms
    return get_fin_terms_cell()

def generate_embeddings(file_path):
    new_namespace = file_path
    clear_namespace(new_namespace, INDEX)

    fin_term_embeds = generate_fin_term_embeds(file_path)  # Assume this function is defined elsewhere
    # sliding_window_embeds = generate_sliding_window_emeds(file_path)  # Assume this function is defined elsewhere
    combined_terms = fin_term_embeds #+ sliding_window_embeds
    chunk_id = 1
    for term in combined_terms:
        term['id'] = str(chunk_id)
        chunk_id +=1
    upsert_chunks(combined_terms, file_path)


def get_formatted_queries(user_query):
    filtered_query = FormatQuestionCall(user_query)
    filtered_query.respond()
    queries_formatted =  ({"text_to_embed": query, "id": str(ix+1)} for ix, query in enumerate(filtered_query.queries))
    print(f'{filtered_query.queries = }')
    return queries_formatted, filtered_query


def embed_query(query):
    temp_client = OpenAI(api_key=EMBED_API_KEY)
    response = temp_client.embeddings.create(
        input=query,
        model="text-embedding-3-small",
    )
    return response.data[0].embedding